import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import random
import math


'''初始化种群'''
def Init(pop, ub, lb, dim):
    '''

    Args:
        pop: 种群数量
        ub: 上边界，维度为[1,dim]
        lb: 下边界，维度为[1,dim]
        dim: 每个个体的维度

    Returns:
        X: 初始化后的种群

    '''
    X = np.zeros([pop, dim, 2])  # 声明空间
    for i in range(pop):
        X[i, :, 0] = np.random.random(dim) * (ub - lb) + lb  # 随机初始化种群
        X[i, :, 1] = np.random.random(dim) * (ub - lb) + lb  # 随机初始化种群

    return X


'''边界约束'''
def Space(X, ub, lb):
    for i in range(len(X)):
        for j in range(2):
            if X[i][j] > ub[0]: # 超出上界
                X[i][j] = ub[0]
            if X[i][j] < lb[0]: # 超出下界
                X[i][j] = ub[0]
    return X


'''HWSN适应度函数，覆盖率计算'''
def fun(X, R, L):
    '''

    Args:
        X: 群体
        R: 感知半径
        L: 监测区域长度

    Returns: 适应度值

    '''

    # 区域覆盖，将区域划分成L*L个单位方格，先计算像素点的联合感知概率，再计算整个覆盖率
    coverNum = 0 # 可以覆盖的像素点个数，初始化为0
    for x in range(L): # 依次计算每一个像素点
        for y in range(L):
            for s in range(len(X)):
                dis = (X[s][0] - x) ** 2 + (X[s][1] - y) ** 2
                if dis < (R**2): # 布尔感知模型
                    coverNum = coverNum + 1
                    break
    result = 1 - (coverNum / (L ** 2)) # 覆盖率=已覆盖的像素点数/所有像素点数，因为是最小优化，所以计算未覆盖率

    return result


'''改进人工蜂鸟算法'''
def AHA(pop, maxIter, dim, L, R, fixed):
    '''

    Args:
        pop: 种群大小
        maxIter: 最大迭代次数
        dim: 维度，即移动节点个数
        L: 上限，方形监测区域边长，下限为0
        R: 感知半径

    Returns:
        bestX: 最优解，画优化后覆盖图用
        bestF: 最优值，表格对比覆盖率用
        hisBestFit: 寻优过程，画迭代图用
        InitSolu: 初始位置，画优化前覆盖图用

    '''

    # region 参数设置与初始化
    hisBestFit = np.zeros(maxIter+1) # 声明历次迭代最优值空间，画迭代图用
    totalNum = dim + len(fixed) # 总节点数量
    lb = np.zeros(dim)  # 下边界，0
    ub = L * np.ones(dim)  # 上边界，传进来的参数L为监测区域长度
    visitTable = np.zeros((pop, pop))  # 声明访问表空间
    for i in range(pop):
        visitTable[i][i] = np.nan  # 初始化访问表，对角线全null，剩余元素全0
    popPos = Init(pop, ub, lb, dim) # PWLCM混沌映射初始化种群
    popFit = np.zeros(pop)  # 声明种群适应度值空间
    solution = np.zeros([pop, totalNum, 2])
    for i in range(pop):
        solution[i] = np.append(popPos[i], fixed, axis=0) # 将固定节点和移动节点组合成一个整体解决方案
        popFit[i] = fun(solution[i], R, L) # 调用fun函数，依次计算每个解决方案的适应度值

    worstF = popFit.max()
    worstIndex = np.argmax(popFit)
    initSolu = solution[worstIndex] # 记录初始解
    hisBestFit[0] = 1 - worstF  # 记录初始覆盖率
    print('初始群体覆盖率为：', 1 - worstF)
    bestF = popFit.min()  # 最优适应度值
    index = np.argmin(popFit)  # 最优个体的索引
    bestX = popPos[index]  # 最优个体
    # endregion

    # 进入迭代
    for t in range(maxIter):
        DirectVector = np.random.randint(-1, 2, (pop, dim, 2))  # 随机轴向飞行或对角飞行
        for po in range(pop):
            for di in range(dim):
                if DirectVector[po][di][0] == 0 and DirectVector[po][di][1] == 0:
                    DirectVector[po,di] = [-1, 1] # 将所有不飞行[0,0]替换
        # endregion
        for i in range(pop):
            matrix = np.random.randn(dim, 2)  # 符合高斯分布的随机矩阵
            if np.random.random() < 0.5:
                # region 引导觅食
                # 确定候选食物源
                maxUnvisitedTime = np.nanmax(visitTable[i, :]) # 第i行最高访问级别，nanmax可以忽略掉null值
                targetFoodIndex = np.nanargmax(visitTable[i, :])
                sumIndex = np.where(visitTable[i, :] == maxUnvisitedTime)
                if np.size(sumIndex) > 1: # 如果同时有两只及以上的蜂鸟有最高访问级别
                    Ind = np.argmin(popFit[sumIndex]) # 进一步对比适应度值，最优的个体为目标食物源
                    targetFoodIndex = sumIndex[0][Ind]
                # 根据候选食物源更新位置，计算新位置的适应度值
                newPopPos = popPos[targetFoodIndex] + matrix * DirectVector[i] * (popPos[i] - popPos[targetFoodIndex])
                newPopPos = Space(newPopPos, ub, lb)
                newPopPos = np.append(newPopPos, fixed, axis=0)
                newPopFit = fun(newPopPos, R, L)
                # 如果新位置更优，更新位置和访问表的第i行第i列
                if newPopFit < popFit[i]:
                    popFit[i] = newPopFit # 更新适应度值
                    popPos[i] = newPopPos[0:dim,:] # 更新位置
                    visitTable[i, :] = visitTable[i, :] + 1 # 更新访问表第i行，访问等级+1
                    visitTable[i, targetFoodIndex] = 0 # 目标食物源更新为0，表示刚访问过
                    visitTable[:, i] = np.nanmax(visitTable, axis=1) + 1 # 更新访问表第i列，为对应行的最高优先级+1
                    visitTable[i, i] = np.nan # 访问表对角线保持null
                # 如果原来的位置更优，只更新访问表的第i行
                else:
                    visitTable[i, :] = visitTable[i, :] + 1
                    visitTable[i, targetFoodIndex] = 0
                # endregion
            else:
                # region 领地觅食
                newPopPos = popPos[i] + matrix * DirectVector[i] * popPos[i] # 注意这儿有个常量
                newPopPos = Space(newPopPos, ub, lb)
                newPopPos = np.append(newPopPos, fixed, axis=0)
                newPopFit = fun(newPopPos, R, L)
                # 如果新位置更优，更新位置和访问表的第i行第i列
                if newPopFit < popFit[i]:
                    popFit[i] = newPopFit
                    popPos[i] = newPopPos[0:dim, :]
                    visitTable[i, :] = visitTable[i, :] + 1
                    visitTable[:, i] = np.nanmax(visitTable, axis=1) + 1
                    visitTable[i, i] = np.nan
                # 如果原来的位置更优，只更新访问表的第i行
                else:
                    visitTable[i, :] = visitTable[i, :] + 1
                # endregion

        # region 迁移觅食
        M = t % (2 * pop)  # 迁移系数
        if M == 0:
            MigrationIndex = np.argmax(popFit)  # 找到最差蜂鸟
            popPos[MigrationIndex, :, 0] = np.random.random(dim) * (ub - lb) + lb  # 更新位置
            popPos[MigrationIndex, :, 1] = np.random.random(dim) * (ub - lb) + lb  # 更新位置
            popPos[MigrationIndex] = Space(popPos[MigrationIndex], ub, lb)  # 边界约束
            newPopPos = np.append(popPos[MigrationIndex], fixed, axis=0)  # 拼接固定节点
            popFit[MigrationIndex] = fun(newPopPos, R, L)  # 计算适应度值
            visitTable[MigrationIndex, :] = visitTable[MigrationIndex, :] + 1  # 更新访问表行
            visitTable[:, MigrationIndex] = np.nanmax(visitTable, axis=1) + 1  # 更新列
            visitTable[MigrationIndex, MigrationIndex] = np.nan  # 对角线保持null
        # endregion


        # region 更新最优解与最优值
        if popFit.min() <= bestF:
            bestF = popFit.min()
            index = np.argsort(popFit)[0]
            bestX = np.append(popPos[index], fixed, axis=0)
        print('第',t+1,'次迭代，最优值为：', 1-bestF)
        hisBestFit[t+1] = 1 - bestF
        # endregion


    return bestX, 1-bestF, hisBestFit, initSolu


'''主函数'''
def main():

    workbook = openpyxl.Workbook() # 创建活动工作薄
    worksheet = workbook.active # 获取活动工作表， 默认就是第一个工作表
    worksheet.title = "sheet1" # 工作表命名
    worksheet.cell(1, 1, '最优值')
    worksheet.cell(1, 2, '迭代数据')

    # 设置参数
    PopNum = 30 # 种群数量
    MaxIterNum = 2 # 最大迭代次数
    ExpNumber = 1 # 实验次数
    TotalNum = 40 # 总节点数
    Length = 50 # 边界长
    Radius = 5# 感知半径
    Fixed = [[5, 5], [5, 10], [5, 10],
       [5, 15], [5, 20], [5,25], [5, 30],
       [5, 35], [5, 40], [5,45]]
    MobiledNum = TotalNum - len(Fixed) # 剩余移动节点个数即维度

    for e in range(ExpNumber):
        print('-----第', e + 1,'次实验-----')
        bestX, bestF, hisBestFit, InitX = AHA(PopNum, MaxIterNum, MobiledNum, Length, Radius, Fixed)
        worksheet.cell(e + 2, 1, float(bestF))

       
        plt.figure(ExpNumber*2+1)
        # 黑色固定节点坐标初始化
        x1 = []
        y1 = []
        # 红色移动节点坐标初始化
        x2 = []
        y2 = []
        # 把数据分成两部分
        for i in range(TotalNum):
            [a, b] = InitX[i]
            if i < MobiledNum:
                x2.append(a)
                y2.append(b)
            else:
                x1.append(a)
                y1.append(b)
        # 设置用于画圆的角度
        theta = np.arange(0, 2 * np.pi, 0.01)
        # 创建图形对象
        fig, ax = plt.subplots(figsize=(7, 7))
        # 设置右侧和上侧均有刻度
        ax.tick_params(axis='both', which='both', direction='in')
        ax.xaxis.set_ticks_position('both')
        ax.yaxis.set_ticks_position('both')
        # 设置字体
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        # 设置标题和坐标轴显示
        plt.title('Coverage results before optimization', fontsize=20)
        plt.xlabel('X(m)', fontsize=20)
        plt.ylabel('Y(m)', fontsize=20)
        plt.xlim(0, Length)
        plt.ylim(0, Length)
        plt.xticks(np.arange(0, Length + 1, 5), fontsize=14)
        plt.yticks(np.arange(0, Length + 1, 5), fontsize=14)
        # 画固定节点
        for i in range(len(x1)):
            a = x1[i]
            b = y1[i]
            cur_x = a + Radius * np.cos(theta)
            cur_y = b + Radius * np.sin(theta)
            plt.scatter(a, b, marker='+', color='black')
            if i == 0:
                plt.plot(cur_x, cur_y, color='black', label='Fixed node')
            else:
                plt.plot(cur_x, cur_y, color='black')
        # 画移动节点
        for i in range(len(x2)):
            a = x2[i]
            b = y2[i]
            cur_x = a + Radius * np.cos(theta)
            cur_y = b + Radius * np.sin(theta)
            plt.scatter(a, b, marker='.', color='red')
            if i == 0:
                plt.plot(cur_x, cur_y, color='red', label='Mobile node')
            else:
                plt.plot(cur_x, cur_y, color='red')
        plt.legend(loc='upper right', fontsize=16)
        # 保存文件，显示图像，dpi调整图像清晰度，bbox使图像可以完整输出
        plt.savefig("a.png", dpi=750, bbox_inches='tight')
        plt.clf()


        # region 画优化后覆盖图
        plt.figure(ExpNumber+1)
        # 黑色固定节点坐标初始化
        x1 = []
        y1 = []
        # 红色移动节点坐标初始化
        x2 = []
        y2 = []
        # 把数据分成两部分
        for i in range(TotalNum):
            [a, b] = bestX[i]
            if i < MobiledNum:
                x2.append(a)
                y2.append(b)
            else:
                x1.append(a)
                y1.append(b)
        # 设置用于画圆的角度
        theta = np.arange(0, 2 * np.pi, 0.01)
        # 创建图形对象
        fig, ax = plt.subplots(figsize=(7, 7))
        # 设置右侧和上侧均有刻度
        ax.tick_params(axis='both', which='both', direction='in')
        ax.xaxis.set_ticks_position('both')
        ax.yaxis.set_ticks_position('both')
        # 设置字体
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        # 设置标题和坐标轴显示
        plt.title('Coverage results after AHA optimization', fontsize=20)
        plt.xlabel('X(m)', fontsize=20)
        plt.ylabel('Y(m)', fontsize=20)
        plt.xlim(0, Length)
        plt.ylim(0, Length)
        plt.xticks(np.arange(0, Length + 1, 5), fontsize=14)
        plt.yticks(np.arange(0, Length + 1, 5), fontsize=14)
        # 画固定节点
        for i in range(len(x1)):
            a = x1[i]
            b = y1[i]
            cur_x = a + Radius * np.cos(theta)
            cur_y = b + Radius * np.sin(theta)
            plt.scatter(a, b, marker='+', color='black')
            if i == 0:
                plt.plot(cur_x, cur_y, color='black', label='Fixed node')
            else:
                plt.plot(cur_x, cur_y, color='black')
        # 画移动节点
        for i in range(len(x2)):
            a = x2[i]
            b = y2[i]
            cur_x = a + Radius * np.cos(theta)
            cur_y = b + Radius * np.sin(theta)
            plt.scatter(a, b, marker='.', color='red')
            if i == 0:
                plt.plot(cur_x, cur_y, color='red', label='Mobile node')
            else:
                plt.plot(cur_x, cur_y, color='red')
        # 保存文件，显示图像
        plt.legend(loc='upper right', fontsize=16)
        # 保存文件，显示图像，dpi调整图像清晰度，bbox使图像可以完整输出
        plt.savefig("a.png", dpi=750, bbox_inches='tight')
        plt.clf()
        # endregion

        # 记录最后一次的迭代数据
        for t in range(MaxIterNum+1):
            worksheet.cell(t + 2, e + 2, hisBestFit[t])

    workbook.save(filename="workspace.xlsx")


if __name__ == "__main__":
    main()